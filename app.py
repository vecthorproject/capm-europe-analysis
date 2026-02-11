import streamlit as st
import pandas as pd
import numpy as np
import yfinance as yf
import io 
import requests 
import datetime
from openpyxl.styles import Alignment, Font, NamedStyle

# =========================
# CONFIGURAZIONE APP
# =========================
st.set_page_config(page_title="Analisi Beta Pro (Ultimate)", layout="wide", page_icon="🏛️")

# =========================
# GESTIONE STATO
# =========================
if 'portfolio' not in st.session_state: 
    st.session_state['portfolio'] = {} 

# =========================
# HELPERS (Mappature)
# =========================
def map_industry_to_ateco(industry):
    mapping = {
        "Banks—Diversified": "64.19 (Banche)", "Auto Manufacturers": "29.10 (Auto)",
        "Luxury Goods": "47.71 (Lusso)", "Utilities—Renewable": "35.11 (Energia)",
        "Oil & Gas Integrated": "06.10 (Oil&Gas)", "Semiconductors": "26.11 (Chip)",
        "Aerospace & Defense": "30.30 (Aerospaziale)", "Insurance—Diversified": "65.12 (Assicurazioni)",
        "Beverages—Wineries & Distilleries": "11.02 (Vini)", "Apparel Manufacturing": "14.13 (Abbigliamento)"
    }
    return mapping.get(industry, "N.D.")

def search_yahoo_finance(query):
    if not query or len(query) < 2: return []
    try:
        url = f"https://query2.finance.yahoo.com/v1/finance/search?q={query}"
        headers = {"User-Agent": "Mozilla/5.0"}
        r = requests.get(url, headers=headers, timeout=3)
        data = r.json()
        results = []
        if 'quotes' in data:
            for item in data['quotes']:
                sym = item.get('symbol')
                name = item.get('shortname') or item.get('longname')
                exch = item.get('exchange')
                if sym and name:
                    results.append((f"{name} ({sym}) - {exch}", sym, name))
        return results
    except: return []

# =========================
# 1. SIDEBAR: PANNELLO DI CONTROLLO
# =========================
st.sidebar.title("⚙️ Configurazione Analisi")

# A. Orizzonte Temporale
st.sidebar.subheader("1. Orizzonte & Frequenza")
time_period = st.sidebar.selectbox(
    "Finestra Storica:", 
    ["5 Anni (Lungo Periodo)", "3 Anni (Medio Periodo)", "1 Anno (Breve Periodo)"],
    index=0
)
frequency = st.sidebar.selectbox("Frequenza Dati:", ["Settimanale (Consigliato)", "Mensile"])

# Setup Date
end_date = datetime.date.today()
if "5 Anni" in time_period: delta_weeks = 260
elif "3 Anni" in time_period: delta_weeks = 156
else: delta_weeks = 52
start_date = end_date - datetime.timedelta(weeks=delta_weeks)

# === LOGICA ACCADEMICA CORRETTA ===
# Se Mensile: Scarico "1mo" nativo (No resample, evita errori fine mese)
# Se Settimanale: Scarico "1d" e poi resample a Venerdì (Standard Bloomberg)
if "Mensile" in frequency:
    yf_interval = "1mo"
    resample_rule = None 
else:
    yf_interval = "1d"
    resample_rule = "W-FRI"

st.sidebar.divider()

# B. Parametri CAPM
st.sidebar.subheader("2. Parametri CAPM (Ke)")
rf_input = st.sidebar.number_input("Risk Free Rate (BTP 10Y) %:", value=3.80, step=0.10, format="%.2f") / 100
mrp_input = st.sidebar.number_input("Market Risk Premium (MRP) %:", value=5.50, step=0.10, format="%.2f") / 100

st.sidebar.divider()

# C. Benchmark
st.sidebar.subheader("3. Benchmark di Mercato")
bench_map = {
    "FTSEMIB.MI": "🇮🇹 Italia (FTSE MIB)", 
    "^STOXX50E": "🇪🇺 Europa (Euro Stoxx 50)", 
    "^GSPC": "🇺🇸 USA (S&P 500)",
    "^GDAXI": "🇩🇪 Germania (DAX)",
    "^FCHI": "🇫🇷 Francia (CAC 40)"
}
sel_bench_code = st.sidebar.selectbox("Indice di Riferimento:", list(bench_map.keys()), format_func=lambda x: bench_map[x])

# =========================
# 2. GESTIONE PORTAFOGLIO
# =========================
st.title("💼 Analisi Beta & Struttura Finanziaria")

col_search, col_add = st.columns([3, 1])
with col_search:
    search_q = st.text_input("🔍 Cerca Titolo:", placeholder="Es. Enel, Intesa, Ferrari...")
    search_res = search_yahoo_finance(search_q) if search_q else []
    selected_asset = st.selectbox("Risultati:", search_res, format_func=lambda x: x[0], label_visibility="collapsed") if search_res else None

with col_add:
    st.write("") 
    st.write("") 
    if st.button("➕ Aggiungi", type="primary", use_container_width=True):
        if selected_asset:
            sym, name = selected_asset[1], selected_asset[2]
            if sym not in st.session_state['portfolio']:
                st.session_state['portfolio'][sym] = name
                st.toast(f"✅ {name} aggiunto!", icon="🚀")
            else:
                st.toast(f"⚠️ {name} già presente!", icon="Duplicato")

# Lista Interattiva
if st.session_state['portfolio']:
    st.markdown("### 📋 Portafoglio Attivo")
    for ticker, name in list(st.session_state['portfolio'].items()):
        c1, c2, c3 = st.columns([0.1, 0.7, 0.2])
        c1.write("📌")
        c2.write(f"**{name}** ({ticker})")
        if c3.button("🗑️ Rimuovi", key=f"del_{ticker}"):
            del st.session_state['portfolio'][ticker]
            st.rerun()
    st.markdown("---")
else:
    st.info("Il portafoglio è vuoto. Cerca un titolo per iniziare l'analisi.")

# =========================
# 3. MOTORE DI CALCOLO (Scientifico)
# =========================
def get_financials_robust(ticker_obj):
    """
    Estrae i dati fondamentali per il calcolo di Hamada.
    Usa 'Total Debt' e 'Market Cap' (Equity Value).
    """
    try:
        info = ticker_obj.info
        bs = ticker_obj.balance_sheet
        
        mkt_cap = info.get('marketCap', 0)
        total_debt = 0.0
        
        # 1. Metodo Prioritario: Yahoo Info (Già aggregato)
        if 'totalDebt' in info and info['totalDebt'] is not None:
             total_debt = float(info['totalDebt'])
        
        # 2. Metodo Fallback: Scansione Bilancio
        if total_debt == 0 and not bs.empty:
            if 'Total Debt' in bs.index:
                total_debt = float(bs.loc['Total Debt'].iloc[0])
            elif 'Long Term Debt' in bs.index:
                lt = float(bs.loc['Long Term Debt'].iloc[0])
                st_debt = float(bs.loc['Current Debt'].iloc[0]) if 'Current Debt' in bs.index else 0.0
                total_debt = lt + st_debt
        
        return {
            "Market Cap": mkt_cap,
            "Total Debt": total_debt,
            "Industry": info.get('industry', 'N.D.'),
            "Tax Rate": 0.24 # Aliquota Marginale Standard Italia
        }
    except: return None

def get_market_data(ticker, start, end, interval, rule_resample):
    """
    Scarica i prezzi e calcola i rendimenti.
    Nota: Usa Rendimenti Semplici (Pct Change) coerenti con il CAPM standard.
    """
    try:
        df = yf.download(ticker, start=start, end=end, interval=interval, progress=False)
        if df.empty: return None
        
        if isinstance(df.columns, pd.MultiIndex): 
            df.columns = df.columns.get_level_values(0)
        
        # Resampling solo se necessario (Settimanale da Giornaliero)
        # Se Mensile (1mo), saltiamo questo step per mantenere il dato nativo
        if rule_resample:
            df_final = df.resample(rule_resample).agg({
                'Open': 'first', 'High': 'max', 'Low': 'min', 'Close': 'last', 'Volume': 'sum'
            }).dropna()
        else:
            df_final = df.copy().dropna()

        # ✅ RENDIMENTI SEMPLICI (Standard CAPM/Beta)
        df_final['Var %'] = df_final['Close'].pct_change()
        return df_final
    except: return None

# =========================
# 4. ESECUZIONE & REPORT
# =========================
if st.session_state['portfolio']:
    if st.button("🚀 Avvia Analisi Scientifica", type="primary", use_container_width=True):
        
        results = {}
        with st.spinner(f"Elaborazione dati ({frequency}) in corso..."):
            
            # Scarico Benchmark
            df_bench = get_market_data(sel_bench_code, start_date, end_date, yf_interval, resample_rule)
            
            for t, t_name in st.session_state['portfolio'].items():
                df_asset = get_market_data(t, start_date, end_date, yf_interval, resample_rule)
                fund = get_financials_robust(yf.Ticker(t))
                
                if df_asset is not None and df_bench is not None:
                    # Sincronizzazione Temporale (Intersection)
                    common = df_asset.index.intersection(df_bench.index)
                    
                    # Controllo Osservazioni Minime (es. 2 anni = ~100 settimane)
                    if len(common) > 10: 
                        y = df_asset.loc[common, 'Var %'].dropna()
                        x = df_bench.loc[common, 'Var %'].dropna()
                        idx_clean = y.index.intersection(x.index)
                        
                        # A. Calcolo Beta Levered (Regressione OLS Statistica)
                        # ddof=1 per Varianza Campionaria (Sample Variance)
                        cov = np.cov(y.loc[idx_clean], x.loc[idx_clean], ddof=1)[0][1]
                        var = np.var(x.loc[idx_clean], ddof=1)
                        beta_lev = cov / var if var != 0 else 0
                        
                        # B. Calcolo Beta Unlevered (Formula Hamada)
                        # Bu = Bl / [1 + (1 - Tax) * (D/E)]
                        beta_unlev = beta_lev
                        d_e_ratio = 0.0
                        
                        if fund and fund['Market Cap'] > 0:
                            d_e_ratio = fund['Total Debt'] / fund['Market Cap']
                            beta_unlev = beta_lev / (1 + (1 - fund['Tax Rate']) * d_e_ratio)
                        
                        # C. Costo Equity (CAPM Standard)
                        ke = rf_input + (beta_lev * mrp_input)

                        metrics = {
                            "Beta Lev": beta_lev,
                            "Beta Unlev": beta_unlev,
                            "Ke": ke,
                            "D/E Ratio": d_e_ratio,
                            "Equity": fund['Market Cap'] if fund else 0,
                            "Debt": fund['Total Debt'] if fund else 0,
                            "ATECO": map_industry_to_ateco(fund.get('Industry') if fund else 'N.D.')
                        }
                        
                        results[t] = {
                            "df": df_asset.loc[idx_clean], 
                            "metrics": metrics,
                            "bs": yf.Ticker(t).balance_sheet,
                            "name": t_name
                        }
        
        # OUTPUT A VIDEO
        if results:
            st.success("Analisi completata.")
            st.subheader(f"📊 Sintesi Coefficienti ({frequency})")
            
            summary_rows = []
            for k, v in results.items():
                m = v['metrics']
                summary_rows.append({
                    "Società": v['name'],
                    "Beta Lev (Bl)": f"{m['Beta Lev']:.3f}",   # 3 Decimali
                    "Beta Asset (Bu)": f"{m['Beta Unlev']:.3f}", # 3 Decimali
                    "D/E (Mkt)": f"{m['D/E Ratio']:.3f}",       # 3 Decimali
                    "Ke (Costo Equity)": f"{m['Ke']*100:.2f}%",
                    "Settore": m['ATECO']
                })
            
            st.dataframe(pd.DataFrame(summary_rows), use_container_width=True)
            
            # Note metodologiche
            st.markdown(f"""
            <div style="background-color:#f9f9f9; padding:10px; border-radius:5px; font-size:12px; color:#555;">
            <strong>Nota Metodologica:</strong><br>
            • <strong>Beta Levered:</strong> Calcolato su rendimenti semplici ({time_period}).<br>
            • <strong>D/E Ratio:</strong> <em>Market D/E</em> (Debito Totale / Capitalizzazione di Mercato).<br>
            • <strong>Beta Unlevered:</strong> De-leveraging via formula di Hamada (Tax Rate 24%).
            </div>
            """, unsafe_allow_html=True)

            # GENERATORE EXCEL PERFETTO
            def make_excel_perfect(res_data):
                bio = io.BytesIO()
                with pd.ExcelWriter(bio, engine='openpyxl') as writer:
                    
                    # --- FOGLIO SINTESI ---
                    df_s = pd.DataFrame(summary_rows)
                    # Convertiamo le stringhe in numeri per Excel prima di scrivere
                    # (Rimuoviamo il % per il Ke per farlo trattare come numero)
                    df_s_excel = df_s.copy()
                    
                    df_s.to_excel(writer, sheet_name="Sintesi", index=False, startrow=1)
                    ws = writer.sheets["Sintesi"]
                    
                    # Titolo
                    ws.cell(1, 1, f"REPORT ANALISI RISCHIO ({time_period})").font = Font(bold=True, size=14, color="000080")
                    
                    # Parametri
                    r_p = len(df_s) + 5
                    ws.cell(r_p, 1, "PARAMETRI DI CALCOLO").font = Font(bold=True)
                    ws.cell(r_p+1, 1, f"Risk Free (Rf): {rf_input*100:.2f}%")
                    ws.cell(r_p+2, 1, f"Market Risk Premium (MRP): {mrp_input*100:.2f}%")
                    ws.cell(r_p+3, 1, f"Benchmark: {bench_map[sel_bench_code]}")
                    
                    # Legenda Formule
                    r_f = r_p + 5
                    ws.cell(r_f, 1, "FORMULE APPLICATE").font = Font(bold=True)
                    ws.cell(r_f+1, 1, "Beta Lev: Cov(Ri, Rm)/Var(Rm) (ddof=1)")
                    ws.cell(r_f+2, 1, "Beta Unlev: Bl / [1 + (1-T)*D/E]")
                    ws.cell(r_f+3, 1, "Ke (CAPM): Rf + Bl * MRP")

                    # Valori Grezzi (Trasparenza)
                    curr = r_f + 5
                    for k, v in res_data.items():
                        m = v['metrics']
                        ws.cell(curr, 1, f"Dati {v['name']} ({k}):").font = Font(bold=True)
                        ws.cell(curr+1, 1, f"- Equity (Mkt Cap): {m['Equity']:,.0f} €")
                        ws.cell(curr+2, 1, f"- Debito Totale: {m['Debt']:,.0f} €")
                        ws.cell(curr+3, 1, f"- D/E Ratio Calc: {m['D/E Ratio']:.4f}")
                        curr += 5

                    # --- FOGLI DATI ---
                    for t, d in res_data.items():
                        # Dati Prezzi
                        df_exp = d['df'].copy().reset_index()
                        df_exp.columns = ["Data", "Apertura", "Massimo", "Minimo", "Chiusura", "Volume", "Var %"]
                        df_exp['Data'] = df_exp['Data'].dt.date
                        
                        safe_name = t.replace(".MI", "").replace(".","")[:20]
                        df_exp.to_excel(writer, sheet_name=f"{safe_name}_Dati", index=False)
                        
                        # Bilancio
                        if not d['bs'].empty:
                            d['bs'].to_excel(writer, sheet_name=f"{safe_name}_BS")
                    
                    # --- FORMATTAZIONE AVANZATA (BODYGUARD ###) ---
                    number_style = NamedStyle(name="decimal_3", number_format="0.000")
                    
                    for sheet in writer.sheets:
                        ws_cur = writer.sheets[sheet]
                        for col in ws_cur.columns:
                            col_letter = col[0].column_letter
                            head = str(col[0].value).lower()
                            
                            # 1. Larghezza Colonne
                            if "data" in head or "date" in head:
                                width = 22 # Larghezza fissa per date
                            else:
                                max_l = 0
                                for cell in col:
                                    try:
                                        if cell.value: max_l = max(max_l, len(str(cell.value)))
                                    except: pass
                                width = min(max_l * 1.2 + 2, 50) # Auto-fit con limite
                            ws_cur.column_dimensions[col_letter].width = width
                            
                            # 2. Allineamento e Formato Numerico
                            for cell in col:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                # Se è un numero (escluse date e intestazioni), applica 3 decimali
                                if isinstance(cell.value, (int, float)) and cell.row > 1:
                                    if "Beta" in head or "D/E" in head or "Ratio" in head:
                                        cell.number_format = "0.000"

                return bio.getvalue()

            st.download_button(
                "📥 Scarica Report Excel (Ultimate)", 
                data=make_excel_perfect(results), 
                file_name=f"Analisi_Beta_Ultimate_{datetime.date.today()}.xlsx", 
                type="primary",
                use_container_width=True
            )
